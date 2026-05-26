"""Advanced visual fixtures: conditional formatting, tables, etc."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    CellIsRule,
    FormulaRule,
)
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_cond_format_color_scale(out_dir))
    files.append(_cond_format_data_bars(out_dir))
    files.append(_cond_format_icon_sets(out_dir))
    files.append(_cond_format_highlight(out_dir))
    files.append(_table_banded(out_dir))
    files.append(_table_styles(out_dir))
    return files


def _cond_format_color_scale(out_dir: Path) -> Path:
    """2-color and 3-color scales on numeric ranges."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Scales"

    # Column A: 2-color scale (white to blue)
    ws["A1"] = "2-Color Scale"
    ws["A1"].font = Font(bold=True)
    for i in range(2, 12):
        ws.cell(row=i, column=1, value=(i - 1) * 10)

    ws.conditional_formatting.add(
        "A2:A11",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max", end_color="0000FF",
        ),
    )

    # Column C: 3-color scale (red-yellow-green)
    ws["C1"] = "3-Color Scale"
    ws["C1"].font = Font(bold=True)
    for i in range(2, 12):
        ws.cell(row=i, column=3, value=(i - 1) * 10)

    ws.conditional_formatting.add(
        "C2:C11",
        ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00FF00",
        ),
    )

    path = out_dir / "cond-format-color-scale.xlsx"
    wb.save(path)
    return path


def _cond_format_data_bars(out_dir: Path) -> Path:
    """Data bars on numeric ranges."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Bars"

    ws["A1"] = "Data Bars"
    ws["A1"].font = Font(bold=True)
    values = [10, 25, 40, 55, 70, 85, 100, 60, 30, 15]
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=1, value=v)

    ws.conditional_formatting.add(
        "A2:A11",
        DataBarRule(
            start_type="min",
            end_type="max",
            color="4472C4",
        ),
    )

    path = out_dir / "cond-format-data-bars.xlsx"
    wb.save(path)
    return path


def _cond_format_icon_sets(out_dir: Path) -> Path:
    """Traffic lights, arrows, stars icon sets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Icon Sets"

    # Column A: 3 traffic lights
    ws["A1"] = "Traffic Lights"
    ws["A1"].font = Font(bold=True)
    for i, v in enumerate([10, 30, 50, 70, 90, 20, 60, 80, 40, 100], start=2):
        ws.cell(row=i, column=1, value=v)

    ws.conditional_formatting.add(
        "A2:A11",
        IconSetRule(
            icon_style="3TrafficLights1",
            type="percent",
            values=[0, 33, 67],
        ),
    )

    # Column C: 3 arrows
    ws["C1"] = "Arrows"
    ws["C1"].font = Font(bold=True)
    for i, v in enumerate([10, 30, 50, 70, 90, 20, 60, 80, 40, 100], start=2):
        ws.cell(row=i, column=3, value=v)

    ws.conditional_formatting.add(
        "C2:C11",
        IconSetRule(
            icon_style="3Arrows",
            type="percent",
            values=[0, 33, 67],
        ),
    )

    path = out_dir / "cond-format-icon-sets.xlsx"
    wb.save(path)
    return path


def _cond_format_highlight(out_dir: Path) -> Path:
    """Highlight cells rules: greater than, text contains, duplicates."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Highlight Rules"

    # Data
    ws["A1"] = "Value"
    ws["A1"].font = Font(bold=True)
    values = [5, 15, 25, 35, 45, 15, 55, 25, 65, 75]
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=1, value=v)

    # Greater than 30 → green fill
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.conditional_formatting.add(
        "A2:A11",
        CellIsRule(
            operator="greaterThan",
            formula=["30"],
            fill=green_fill,
        ),
    )

    # Text contains
    ws["C1"] = "Name"
    ws["C1"].font = Font(bold=True)
    names = ["Apple", "Banana", "Apricot", "Cherry", "Avocado", "Blueberry", "Grape", "Peach", "Plum", "Mango"]
    for i, n in enumerate(names, start=2):
        ws.cell(row=i, column=3, value=n)

    # Highlight cells containing "a" (case-insensitive via formula)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.conditional_formatting.add(
        "C2:C11",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("a",C2))'],
            fill=yellow_fill,
        ),
    )

    path = out_dir / "cond-format-highlight.xlsx"
    wb.save(path)
    return path


def _table_banded(out_dir: Path) -> Path:
    """Excel Table with banded rows, header, total row, filter arrows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    data = [
        ["Name", "Department", "Salary"],
        ["Alice", "Engineering", 95000],
        ["Bob", "Marketing", 72000],
        ["Carol", "Engineering", 105000],
        ["Dave", "Sales", 68000],
        ["Eve", "Engineering", 98000],
    ]
    for row in data:
        ws.append(row)

    tab = Table(displayName="EmployeeTable", ref="A1:C6")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    path = out_dir / "table-banded.xlsx"
    wb.save(path)
    return path


def _table_styles(out_dir: Path) -> Path:
    """Tables with different built-in styles."""
    wb = Workbook()

    styles = [
        ("Light", "TableStyleLight1"),
        ("Medium", "TableStyleMedium9"),
        ("Dark", "TableStyleDark1"),
    ]

    for sheet_name, style_name in styles:
        ws = wb.create_sheet(sheet_name)
        data = [
            ["Item", "Count", "Price"],
            ["Widget A", 100, 9.99],
            ["Widget B", 250, 14.50],
            ["Widget C", 75, 22.00],
        ]
        for row in data:
            ws.append(row)

        tab = Table(displayName=f"Table{sheet_name}", ref="A1:C4")
        tab.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showRowStripes=True,
        )
        ws.add_table(tab)

    # Remove default empty sheet
    del wb["Sheet"]

    path = out_dir / "table-styles.xlsx"
    wb.save(path)
    return path
