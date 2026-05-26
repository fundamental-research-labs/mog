"""Behavioral test fixtures with companion test sequence definitions."""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate(out_dir: Path) -> list[Path]:
    files = []
    test_sequences = {}

    result, seq = _nav_basic(out_dir)
    files.append(result)
    test_sequences["nav-basic.xlsx"] = seq

    result, seq = _nav_merged_cells(out_dir)
    files.append(result)
    test_sequences["nav-merged-cells.xlsx"] = seq

    result, seq = _nav_ctrl_arrow(out_dir)
    files.append(result)
    test_sequences["nav-ctrl-arrow.xlsx"] = seq

    result, seq = _nav_ctrl_home_end(out_dir)
    files.append(result)
    test_sequences["nav-ctrl-home-end.xlsx"] = seq

    result, seq = _select_shift_arrow(out_dir)
    files.append(result)
    test_sequences["select-shift-arrow.xlsx"] = seq

    result, seq = _edit_type_confirm(out_dir)
    files.append(result)
    test_sequences["edit-type-confirm.xlsx"] = seq

    result, seq = _edit_escape_cancel(out_dir)
    files.append(result)
    test_sequences["edit-escape-cancel.xlsx"] = seq

    result, seq = _undo_redo(out_dir)
    files.append(result)
    test_sequences["undo-redo.xlsx"] = seq

    result, seq = _delete_clear(out_dir)
    files.append(result)
    test_sequences["delete-clear.xlsx"] = seq

    result, seq = _clipboard_copy_paste(out_dir)
    files.append(result)
    test_sequences["clipboard-copy-paste.xlsx"] = seq

    result, seq = _sheet_tab_switch(out_dir)
    files.append(result)
    test_sequences["sheet-tab-switch.xlsx"] = seq

    result, seq = _sort_filter(out_dir)
    files.append(result)
    test_sequences["sort-filter.xlsx"] = seq

    result, seq = _fill_handle_series(out_dir)
    files.append(result)
    test_sequences["fill-handle-series.xlsx"] = seq

    # Write test sequences JSON
    seq_path = out_dir / "test-sequences.json"
    with open(seq_path, "w") as f:
        json.dump(test_sequences, f, indent=2)
    files.append(seq_path)

    return files


def _nav_basic(out_dir: Path) -> tuple[Path, dict]:
    """5x5 data grid for arrow key navigation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Navigation"
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=f"{get_column_letter(c)}{r}")

    path = out_dir / "nav-basic.xlsx"
    wb.save(path)

    seq = {
        "description": "Arrow key navigation across a 5x5 data grid",
        "start_cell": "A1",
        "steps": [
            {"action": "key", "key": "ArrowRight", "repeat": 3, "expect_active": "D1"},
            {"action": "key", "key": "ArrowDown", "repeat": 2, "expect_active": "D3"},
            {"action": "key", "key": "ArrowLeft", "repeat": 3, "expect_active": "A3"},
            {"action": "key", "key": "ArrowUp", "repeat": 2, "expect_active": "A1"},
        ],
    }
    return path, seq


def _nav_merged_cells(out_dir: Path) -> tuple[Path, dict]:
    """Merged cell regions for navigation testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged Nav"

    # Fill grid
    for r in range(1, 8):
        for c in range(1, 8):
            ws.cell(row=r, column=c, value=f"{get_column_letter(c)}{r}")

    # Horizontal merge B2:D2
    ws.merge_cells("B2:D2")
    ws["B2"] = "Merged H"

    # Vertical merge F3:F5
    ws.merge_cells("F3:F5")
    ws["F3"] = "Merged V"

    path = out_dir / "nav-merged-cells.xlsx"
    wb.save(path)

    seq = {
        "description": "Navigate into and across merged cell regions",
        "start_cell": "A2",
        "steps": [
            {"action": "key", "key": "ArrowRight", "expect_active": "B2", "note": "Enter horizontal merge"},
            {"action": "key", "key": "ArrowRight", "expect_active": "E2", "note": "Exit merge, skip to E2"},
        ],
    }
    return path, seq


def _nav_ctrl_arrow(out_dir: Path) -> tuple[Path, dict]:
    """Sparse data grid for Ctrl+Arrow boundary jumps."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ctrl Arrow"

    # Column A: filled 1-10, gap at 11-15, filled 16-20
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=f"Data {r}")
    for r in range(16, 21):
        ws.cell(row=r, column=1, value=f"Data {r}")

    # Column C: sparse — filled at rows 1, 5, 10
    ws["C1"] = "Start"
    ws["C5"] = "Middle"
    ws["C10"] = "End"

    path = out_dir / "nav-ctrl-arrow.xlsx"
    wb.save(path)

    seq = {
        "description": "Ctrl+Arrow jumps to data boundaries in sparse grid",
        "start_cell": "A1",
        "steps": [
            {"action": "key", "key": "Ctrl+ArrowDown", "expect_active": "A10", "note": "Jump to last filled before gap"},
            {"action": "key", "key": "Ctrl+ArrowDown", "expect_active": "A16", "note": "Jump to first filled after gap"},
            {"action": "key", "key": "Ctrl+ArrowDown", "expect_active": "A20", "note": "Jump to last filled"},
        ],
    }
    return path, seq


def _nav_ctrl_home_end(out_dir: Path) -> tuple[Path, dict]:
    """Data in A1:E50 for Ctrl+Home/End testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ctrl Home End"

    for r in range(1, 51):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=r * c)

    path = out_dir / "nav-ctrl-home-end.xlsx"
    wb.save(path)

    seq = {
        "description": "Ctrl+Home and Ctrl+End navigation",
        "start_cell": "C25",
        "steps": [
            {"action": "key", "key": "Ctrl+End", "expect_active": "E50"},
            {"action": "key", "key": "Ctrl+Home", "expect_active": "A1"},
        ],
    }
    return path, seq


def _select_shift_arrow(out_dir: Path) -> tuple[Path, dict]:
    """Grid for Shift+Arrow selection extension."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Selection"
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=f"{get_column_letter(c)}{r}")

    path = out_dir / "select-shift-arrow.xlsx"
    wb.save(path)

    seq = {
        "description": "Shift+Arrow extends selection",
        "start_cell": "B2",
        "steps": [
            {"action": "key", "key": "Shift+ArrowRight", "repeat": 2, "expect_selection": "B2:D2"},
            {"action": "key", "key": "Shift+ArrowDown", "repeat": 2, "expect_selection": "B2:D4"},
        ],
    }
    return path, seq


def _edit_type_confirm(out_dir: Path) -> tuple[Path, dict]:
    """Empty grid for typing and Enter confirmation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Editing"
    # Leave cells empty for editing test

    path = out_dir / "edit-type-confirm.xlsx"
    wb.save(path)

    seq = {
        "description": "Type value and press Enter to confirm",
        "start_cell": "A1",
        "steps": [
            {"action": "type", "text": "hello", "then": "Enter"},
            {"expect_value": {"A1": "hello"}, "expect_active": "A2"},
        ],
    }
    return path, seq


def _edit_escape_cancel(out_dir: Path) -> tuple[Path, dict]:
    """Cell with existing value for edit cancellation test."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cancel Edit"
    ws["A1"] = "Original"

    path = out_dir / "edit-escape-cancel.xlsx"
    wb.save(path)

    seq = {
        "description": "Escape cancels in-progress edit",
        "start_cell": "A1",
        "steps": [
            {"action": "type", "text": "Changed"},
            {"action": "key", "key": "Escape"},
            {"expect_value": {"A1": "Original"}, "note": "Value should be unchanged"},
        ],
    }
    return path, seq


def _undo_redo(out_dir: Path) -> tuple[Path, dict]:
    """Cell for undo/redo testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Undo Redo"
    ws["A1"] = "Original"

    path = out_dir / "undo-redo.xlsx"
    wb.save(path)

    seq = {
        "description": "Ctrl+Z undo and Ctrl+Y redo",
        "start_cell": "A1",
        "steps": [
            {"action": "type", "text": "Changed", "then": "Enter"},
            {"expect_value": {"A1": "Changed"}},
            {"action": "key", "key": "Ctrl+Z"},
            {"expect_value": {"A1": "Original"}, "note": "Undo restores original"},
            {"action": "key", "key": "Ctrl+Y"},
            {"expect_value": {"A1": "Changed"}, "note": "Redo restores edit"},
        ],
    }
    return path, seq


def _delete_clear(out_dir: Path) -> tuple[Path, dict]:
    """Cell with value and formatting for Delete key test."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Delete"
    ws["A1"] = "DeleteMe"
    ws["A1"].font = Font(bold=True, color="FF0000")
    ws["A1"].fill = PatternFill(start_color="FFFF00", fill_type="solid")

    path = out_dir / "delete-clear.xlsx"
    wb.save(path)

    seq = {
        "description": "Delete key clears value but preserves formatting",
        "start_cell": "A1",
        "steps": [
            {"action": "key", "key": "Delete"},
            {"expect_value": {"A1": ""}, "note": "Value cleared, formatting preserved"},
        ],
    }
    return path, seq


def _clipboard_copy_paste(out_dir: Path) -> tuple[Path, dict]:
    """Source cell with value and formatting for copy/paste."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clipboard"
    ws["A1"] = "CopyMe"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill(start_color="00FF00", fill_type="solid")

    path = out_dir / "clipboard-copy-paste.xlsx"
    wb.save(path)

    seq = {
        "description": "Ctrl+C copy, Ctrl+V paste with value and formatting",
        "start_cell": "A1",
        "steps": [
            {"action": "key", "key": "Ctrl+C"},
            {"action": "click", "cell": "C3"},
            {"action": "key", "key": "Ctrl+V"},
            {"expect_value": {"C3": "CopyMe"}, "note": "Value and formatting should transfer"},
        ],
    }
    return path, seq


def _sheet_tab_switch(out_dir: Path) -> tuple[Path, dict]:
    """3 sheets for Ctrl+PageUp/PageDown navigation."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "You are on Sheet1"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "You are on Sheet2"

    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "You are on Sheet3"

    path = out_dir / "sheet-tab-switch.xlsx"
    wb.save(path)

    seq = {
        "description": "Ctrl+PageDown/PageUp switches sheets",
        "start_sheet": "Sheet1",
        "steps": [
            {"action": "key", "key": "Ctrl+PageDown", "expect_sheet": "Sheet2"},
            {"action": "key", "key": "Ctrl+PageDown", "expect_sheet": "Sheet3"},
            {"action": "key", "key": "Ctrl+PageUp", "expect_sheet": "Sheet2"},
        ],
    }
    return path, seq


def _sort_filter(out_dir: Path) -> tuple[Path, dict]:
    """Table with headers for sort/filter testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sort Filter"

    data = [
        ["Name", "Score", "Grade"],
        ["Charlie", 85, "B"],
        ["Alice", 95, "A"],
        ["Eve", 72, "C"],
        ["Bob", 88, "B"],
        ["Dave", 91, "A"],
    ]
    for row in data:
        ws.append(row)

    # Bold header
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = Font(bold=True)

    ws.auto_filter.ref = "A1:C6"

    path = out_dir / "sort-filter.xlsx"
    wb.save(path)

    seq = {
        "description": "Sort by Score ascending",
        "steps": [
            {"action": "sort", "column": "B", "order": "ascending"},
            {"expect_order": ["Eve", "Charlie", "Bob", "Dave", "Alice"]},
        ],
    }
    return path, seq


def _fill_handle_series(out_dir: Path) -> tuple[Path, dict]:
    """Numeric series for fill handle drag."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fill Handle"
    ws["A1"] = 1
    ws["A2"] = 2

    path = out_dir / "fill-handle-series.xlsx"
    wb.save(path)

    seq = {
        "description": "Drag fill handle to extend numeric series",
        "start_selection": "A1:A2",
        "steps": [
            {"action": "fill_drag", "to": "A5"},
            {"expect_value": {"A3": 3, "A4": 4, "A5": 5}},
        ],
    }
    return path, seq
