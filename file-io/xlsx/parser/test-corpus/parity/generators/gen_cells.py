"""Cell-level visual fixtures: formatting, merges, frozen panes, etc."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink


def generate(out_dir: Path) -> list[Path]:
    """Generate all cell-level fixtures. Returns list of created file paths."""
    files = []
    files.append(_basic_formatting(out_dir))
    files.append(_borders(out_dir))
    files.append(_number_formats(out_dir))
    files.append(_merged_cells(out_dir))
    files.append(_frozen_panes(out_dir))
    files.append(_rich_text(out_dir))
    files.append(_hyperlinks(out_dir))
    files.append(_data_validation(out_dir))
    files.append(_text_alignment(out_dir))
    files.append(_column_row_sizing(out_dir))
    files.append(_wide_merge_empty_subcells(out_dir))
    files.append(_spacer_content_dimensions(out_dir))
    files.append(_hello_world(out_dir))
    return files


def _basic_formatting(out_dir: Path) -> Path:
    """Bold, italic, underline, strikethrough, font size, font color, fill color."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formatting"

    ws["A1"] = "Bold"
    ws["A1"].font = Font(bold=True)

    ws["B1"] = "Italic"
    ws["B1"].font = Font(italic=True)

    ws["C1"] = "Underline"
    ws["C1"].font = Font(underline="single")

    ws["D1"] = "Strikethrough"
    ws["D1"].font = Font(strike=True)

    ws["A2"] = "Size 18"
    ws["A2"].font = Font(size=18)

    ws["B2"] = "Red Text"
    ws["B2"].font = Font(color="FF0000")

    ws["C2"] = "Blue Fill"
    ws["C2"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    ws["C2"].font = Font(color="FFFFFF")
    ws["C2"].value = "Blue Fill"

    ws["D2"] = "Bold Red 14pt"
    ws["D2"].font = Font(bold=True, color="FF0000", size=14)

    ws["A3"] = "Arial"
    ws["A3"].font = Font(name="Arial", size=12)

    ws["B3"] = "Courier"
    ws["B3"].font = Font(name="Courier New", size=12)

    path = out_dir / "basic-formatting.xlsx"
    wb.save(path)
    return path


def _borders(out_dir: Path) -> Path:
    """Thin, medium, thick, double, dashed borders + diagonal."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Borders"

    styles = [
        ("Thin", "thin"),
        ("Medium", "medium"),
        ("Thick", "thick"),
        ("Double", "double"),
        ("Dashed", "dashed"),
        ("Dotted", "dotted"),
    ]

    for i, (label, style) in enumerate(styles, start=1):
        cell = ws.cell(row=i, column=1, value=label)
        side = Side(style=style, color="000000")
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    # Diagonal border
    ws["C1"] = "Diagonal Down"
    ws["C1"].border = Border(
        diagonal=Side(style="thin", color="FF0000"),
        diagonalDown=True,
    )
    ws["C2"] = "Diagonal Up"
    ws["C2"].border = Border(
        diagonal=Side(style="thin", color="0000FF"),
        diagonalUp=True,
    )

    path = out_dir / "borders.xlsx"
    wb.save(path)
    return path


def _number_formats(out_dir: Path) -> Path:
    """Currency, percentage, date, time, scientific, custom formats."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Number Formats"

    data = [
        ("Format", "Value", "Display"),
        ("Currency", 1234.56, None),
        ("Percentage", 0.8525, None),
        ("Date", 45000, None),  # Excel serial date
        ("Time", 0.75, None),  # 18:00
        ("Scientific", 123456789, None),
        ("Custom #,##0.00", 1234567.891, None),
        ("Accounting", -500.00, None),
        ("Fraction", 0.333333, None),
    ]

    formats = [
        None,
        "$#,##0.00",
        "0.00%",
        "yyyy-mm-dd",
        "hh:mm:ss",
        "0.00E+00",
        "#,##0.00",
        '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
        "# ?/?",
    ]

    for row_idx, (label, value, _) in enumerate(data, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        if row_idx > 1:
            c = ws.cell(row=row_idx, column=2, value=value)
            if formats[row_idx - 1]:
                c.number_format = formats[row_idx - 1]

    path = out_dir / "number-formats.xlsx"
    wb.save(path)
    return path


def _merged_cells(out_dir: Path) -> Path:
    """Horizontal, vertical, and block merges with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged Cells"

    # Horizontal merge
    ws.merge_cells("A1:C1")
    ws["A1"] = "Horizontal Merge (A1:C1)"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True)

    # Vertical merge
    ws.merge_cells("A3:A6")
    ws["A3"] = "Vertical Merge"
    ws["A3"].alignment = Alignment(vertical="center", text_rotation=90)

    # Block merge
    ws.merge_cells("C3:E5")
    ws["C3"] = "Block Merge (C3:E5)"
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Small merge with border
    ws.merge_cells("G1:H2")
    ws["G1"] = "Bordered Merge"
    ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
    side = Side(style="medium", color="000000")
    ws["G1"].border = Border(left=side, right=side, top=side, bottom=side)

    path = out_dir / "merged-cells.xlsx"
    wb.save(path)
    return path


def _frozen_panes(out_dir: Path) -> Path:
    """Frozen first row, first column, both, and arbitrary position."""
    wb = Workbook()

    # Sheet 1: Frozen first row
    ws1 = wb.active
    ws1.title = "Frozen Row"
    ws1.freeze_panes = "A2"
    ws1["A1"] = "Header A"
    ws1["B1"] = "Header B"
    ws1["C1"] = "Header C"
    ws1["A1"].font = Font(bold=True)
    ws1["B1"].font = Font(bold=True)
    ws1["C1"].font = Font(bold=True)
    for r in range(2, 30):
        for c in range(1, 4):
            ws1.cell(row=r, column=c, value=f"R{r}C{c}")

    # Sheet 2: Frozen first column
    ws2 = wb.create_sheet("Frozen Column")
    ws2.freeze_panes = "B1"
    for r in range(1, 20):
        ws2.cell(row=r, column=1, value=f"Row {r}")
        for c in range(2, 10):
            ws2.cell(row=r, column=c, value=f"Data {r}-{c}")

    # Sheet 3: Frozen both (top-left quadrant)
    ws3 = wb.create_sheet("Frozen Both")
    ws3.freeze_panes = "B2"
    ws3["A1"] = "Corner"
    ws3["A1"].font = Font(bold=True)
    for c in range(2, 10):
        ws3.cell(row=1, column=c, value=f"Col {get_column_letter(c)}")
    for r in range(2, 20):
        ws3.cell(row=r, column=1, value=f"Row {r}")
        for c in range(2, 10):
            ws3.cell(row=r, column=c, value=r * c)

    # Sheet 4: Arbitrary freeze at C5
    ws4 = wb.create_sheet("Frozen C5")
    ws4.freeze_panes = "C5"
    for r in range(1, 30):
        for c in range(1, 10):
            ws4.cell(row=r, column=c, value=f"{get_column_letter(c)}{r}")

    path = out_dir / "frozen-panes.xlsx"
    wb.save(path)
    return path


def _rich_text(out_dir: Path) -> Path:
    """Mixed formatting runs within single cells using openpyxl InlineString."""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    wb = Workbook()
    ws = wb.active
    ws.title = "Rich Text"

    # Cell with bold + italic + colored text in one cell
    rt1 = CellRichText(
        TextBlock(InlineFont(b=True), "Bold "),
        TextBlock(InlineFont(i=True), "Italic "),
        TextBlock(InlineFont(color="FF0000"), "Red"),
    )
    ws["A1"] = rt1

    # Mixed sizes
    rt2 = CellRichText(
        TextBlock(InlineFont(sz=10), "Small "),
        TextBlock(InlineFont(sz=18, b=True), "BIG "),
        TextBlock(InlineFont(sz=10), "Small"),
    )
    ws["A2"] = rt2

    # Underline + strikethrough mix
    rt3 = CellRichText(
        TextBlock(InlineFont(u="single"), "Underlined "),
        "Normal ",
        TextBlock(InlineFont(strike=True), "Struck"),
    )
    ws["A3"] = rt3

    path = out_dir / "rich-text.xlsx"
    wb.save(path)
    return path


def _hyperlinks(out_dir: Path) -> Path:
    """URL, email, and internal sheet reference hyperlinks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hyperlinks"

    # URL hyperlink
    ws["A1"] = "Visit Example.com"
    ws["A1"].hyperlink = "https://example.com"
    ws["A1"].font = Font(color="0000FF", underline="single")

    # Email hyperlink
    ws["A2"] = "Send Email"
    ws["A2"].hyperlink = "mailto:test@example.com"
    ws["A2"].font = Font(color="0000FF", underline="single")

    # Internal reference
    ws2 = wb.create_sheet("Target")
    ws2["A1"] = "You arrived here!"

    ws["A3"] = "Go to Target Sheet"
    ws["A3"].hyperlink = "#Target!A1"
    ws["A3"].font = Font(color="0000FF", underline="single")

    path = out_dir / "hyperlinks.xlsx"
    wb.save(path)
    return path


def _data_validation(out_dir: Path) -> Path:
    """Dropdown list, number range, date range validations."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Validation"

    # Labels
    ws["A1"] = "Dropdown"
    ws["A2"] = "Number (1-100)"
    ws["A3"] = "Text Length (max 10)"

    # Dropdown list
    dv_list = DataValidation(
        type="list",
        formula1='"Apple,Banana,Cherry,Date,Elderberry"',
        allow_blank=True,
    )
    dv_list.prompt = "Pick a fruit"
    dv_list.promptTitle = "Fruit Selection"
    dv_list.error = "Must be a valid fruit"
    dv_list.errorTitle = "Invalid"
    ws.add_data_validation(dv_list)
    dv_list.add("B1")
    ws["B1"] = "Apple"

    # Number range
    dv_num = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="100",
    )
    dv_num.prompt = "Enter a number 1-100"
    ws.add_data_validation(dv_num)
    dv_num.add("B2")
    ws["B2"] = 50

    # Text length
    dv_text = DataValidation(
        type="textLength",
        operator="lessThanOrEqual",
        formula1="10",
    )
    dv_text.prompt = "Max 10 characters"
    ws.add_data_validation(dv_text)
    dv_text.add("B3")
    ws["B3"] = "Short"

    path = out_dir / "data-validation.xlsx"
    wb.save(path)
    return path


def _text_alignment(out_dir: Path) -> Path:
    """Horizontal, vertical alignment, rotation, wrap, shrink, indent."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alignment"

    # Set column width for visibility
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Horizontal alignments
    for i, h_align in enumerate(["left", "center", "right", "fill", "justify"], start=1):
        c = ws.cell(row=i, column=1, value=f"H: {h_align}")
        c.alignment = Alignment(horizontal=h_align)

    # Vertical alignments
    for i, v_align in enumerate(["top", "center", "bottom"], start=1):
        ws.row_dimensions[i].height = 40
        c = ws.cell(row=i, column=2, value=f"V: {v_align}")
        c.alignment = Alignment(vertical=v_align)

    # Text rotation
    ws.row_dimensions[1].height = 60
    c = ws.cell(row=1, column=3, value="45° rotation")
    c.alignment = Alignment(text_rotation=45)

    c = ws.cell(row=2, column=3, value="90° rotation")
    c.alignment = Alignment(text_rotation=90)

    # Wrap text
    c = ws.cell(row=4, column=3, value="This is a long text that should wrap within the cell boundaries")
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 40

    # Shrink to fit
    c = ws.cell(row=5, column=3, value="Shrink to fit this long text")
    c.alignment = Alignment(shrink_to_fit=True)

    # Indent
    c = ws.cell(row=6, column=1, value="Indent level 1")
    c.alignment = Alignment(indent=1)
    c = ws.cell(row=7, column=1, value="Indent level 3")
    c.alignment = Alignment(indent=3)

    path = out_dir / "text-alignment.xlsx"
    wb.save(path)
    return path


def _column_row_sizing(out_dir: Path) -> Path:
    """Wide/tall/hidden columns and rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sizing"

    # Normal column
    ws["A1"] = "Normal Width"

    # Wide column
    ws.column_dimensions["B"].width = 40
    ws["B1"] = "Wide Column (40)"

    # Narrow column
    ws.column_dimensions["C"].width = 5
    ws["C1"] = "Narrow"

    # Hidden column
    ws.column_dimensions["D"].hidden = True
    ws["D1"] = "Hidden Column"

    ws["E1"] = "After Hidden"

    # Tall row
    ws.row_dimensions[3].height = 60
    ws["A3"] = "Tall Row (60pt)"

    # Hidden row
    ws.row_dimensions[4].hidden = True
    ws["A4"] = "Hidden Row"
    ws["A5"] = "After Hidden Row"

    # Auto-fit content (just put long text)
    ws["A7"] = "Short"
    ws["B7"] = "This cell has much longer content for auto-fit testing"

    path = out_dir / "column-row-sizing.xlsx"
    wb.save(path)
    return path


def _wide_merge_empty_subcells(out_dir: Path) -> Path:
    """Wide merge where only the anchor cell has a value; sub-cells are empty."""
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Liability Calculations"
    ws.merge_cells("A1:P1")

    path = out_dir / "wide-merge-empty-subcells.xlsx"
    wb.save(path)
    return path


def _hello_world(out_dir: Path) -> Path:
    """Minimal XLSX with 'hello world' in A1."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello world"

    path = out_dir / "hello-world.xlsx"
    wb.save(path)
    return path


def _spacer_content_dimensions(out_dir: Path) -> Path:
    """Mixed spacer/content rows and columns with explicit dimensions."""
    wb = Workbook()
    ws = wb.active

    # Sheet-level defaults
    ws.sheet_format.defaultColWidth = 9.140625
    ws.sheet_format.defaultRowHeight = 15

    # Column widths (setting width auto-sets customWidth=True)
    for letter, width in [("A", 3.71), ("B", 17.71), ("C", 2.28),
                          ("D", 17.71), ("E", 2.28), ("F", 17.71)]:
        ws.column_dimensions[letter].width = width

    # Row heights (setting height auto-sets customHeight=True)
    for row, height in [(1, 69.75), (2, 60), (3, 6), (4, 15), (5, 79.5)]:
        ws.row_dimensions[row].height = height

    # Label cells
    ws["A1"] = "Header"
    ws["A2"] = "Content"
    ws["A3"] = "Spacer"
    ws["B1"] = "Wide Col"

    path = out_dir / "spacer-content-dimensions.xlsx"
    wb.save(path)
    return path
