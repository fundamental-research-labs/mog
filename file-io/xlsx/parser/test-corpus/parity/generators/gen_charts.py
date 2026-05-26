"""Chart fixtures: bar, line, pie, scatter, area, combo."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    ScatterChart,
    AreaChart,
    Reference,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_chart_bar(out_dir))
    files.append(_chart_bar_stacked(out_dir))
    files.append(_chart_line(out_dir))
    files.append(_chart_pie(out_dir))
    files.append(_chart_scatter(out_dir))
    files.append(_chart_area(out_dir))
    files.append(_chart_combo(out_dir))
    files.append(_chart_with_legend(out_dir))
    files.append(_chart_mini(out_dir))
    return files


def _sample_data(ws):
    """Write common sample data to worksheet."""
    data = [
        ["Category", "Series 1", "Series 2"],
        ["Q1", 10, 15],
        ["Q2", 25, 18],
        ["Q3", 17, 22],
        ["Q4", 30, 28],
    ]
    for row in data:
        ws.append(row)
    return len(data)


def _chart_bar(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    chart = BarChart()
    chart.title = "Quarterly Revenue"
    chart.y_axis.title = "Revenue ($K)"
    chart.x_axis.title = "Quarter"
    chart.style = 10

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=nrows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E2")

    path = out_dir / "chart-bar.xlsx"
    wb.save(path)
    return path


def _chart_bar_stacked(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    chart = BarChart()
    chart.title = "Stacked Bar"
    chart.grouping = "stacked"

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=nrows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E2")

    path = out_dir / "chart-bar-stacked.xlsx"
    wb.save(path)
    return path


def _chart_line(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    chart = LineChart()
    chart.title = "Trend Line"
    chart.y_axis.title = "Value"

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=nrows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Add markers
    for series in chart.series:
        series.graphicalProperties.line.width = 25000  # 2pt

    ws.add_chart(chart, "E2")

    path = out_dir / "chart-line.xlsx"
    wb.save(path)
    return path


def _chart_pie(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    data = [
        ["Category", "Value"],
        ["Desktop", 45],
        ["Mobile", 30],
        ["Tablet", 15],
        ["Other", 10],
    ]
    for row in data:
        ws.append(row)

    chart = PieChart()
    chart.title = "Device Distribution"

    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    vals = Reference(ws, min_col=2, min_row=1, max_row=5)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    # Show percentage labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True

    ws.add_chart(chart, "D2")

    path = out_dir / "chart-pie.xlsx"
    wb.save(path)
    return path


def _chart_scatter(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    data = [
        ["X", "Y"],
        [1, 2.1],
        [2, 4.3],
        [3, 5.8],
        [4, 8.2],
        [5, 10.1],
    ]
    for row in data:
        ws.append(row)

    chart = ScatterChart()
    chart.title = "Scatter Plot"
    chart.x_axis.title = "X"
    chart.y_axis.title = "Y"

    x_vals = Reference(ws, min_col=1, min_row=2, max_row=6)
    y_vals = Reference(ws, min_col=2, min_row=2, max_row=6)
    series = chart.series
    from openpyxl.chart import Series
    s = Series(y_vals, x_vals, title="Data")
    chart.series.append(s)

    # Add trendline
    from openpyxl.chart.trendline import Trendline
    s.trendline = Trendline(trendlineType="linear")

    ws.add_chart(chart, "D2")

    path = out_dir / "chart-scatter.xlsx"
    wb.save(path)
    return path


def _chart_area(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    chart = AreaChart()
    chart.title = "Area Chart"

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=nrows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E2")

    path = out_dir / "chart-area.xlsx"
    wb.save(path)
    return path


def _chart_combo(out_dir: Path) -> Path:
    """Bar + line combination chart."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    # Bar chart (primary)
    bar = BarChart()
    bar.title = "Combo: Bar + Line"

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    bar_data = Reference(ws, min_col=2, min_row=1, max_row=nrows)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(cats)

    # Line chart (secondary axis)
    line = LineChart()
    line_data = Reference(ws, min_col=3, min_row=1, max_row=nrows)
    line.add_data(line_data, titles_from_data=True)
    line.y_axis.axId = 200

    # Combine
    bar += line

    ws.add_chart(bar, "E2")

    path = out_dir / "chart-combo.xlsx"
    wb.save(path)
    return path


def _chart_with_legend(out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    nrows = _sample_data(ws)

    chart = BarChart()
    chart.title = "Chart with Legend"

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=nrows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Legend at bottom
    from openpyxl.chart.layout import Layout, ManualLayout
    chart.legend.position = "b"

    ws.add_chart(chart, "E2")

    path = out_dir / "chart-with-legend.xlsx"
    wb.save(path)
    return path


def _chart_mini(out_dir: Path) -> Path:
    """Small chart for inline/compact rendering."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    data = [["Month", "Val"], ["Jan", 5], ["Feb", 8], ["Mar", 3]]
    for row in data:
        ws.append(row)

    chart = BarChart()
    chart.title = "Mini"
    chart.width = 8
    chart.height = 6

    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    vals = Reference(ws, min_col=2, min_row=1, max_row=4)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "D1")

    path = out_dir / "chart-mini.xlsx"
    wb.save(path)
    return path
