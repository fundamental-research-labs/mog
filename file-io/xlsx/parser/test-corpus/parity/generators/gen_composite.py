"""Kitchen sink composite fixture combining multiple features."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage
from io import BytesIO


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_kitchen_sink(out_dir))
    return files


def _kitchen_sink(out_dir: Path) -> Path:
    """Multi-sheet composite file exercising many features."""
    wb = Workbook()

    # ── Sheet 1: Formatting ──
    ws1 = wb.active
    ws1.title = "Formatting"

    ws1["A1"] = "Mixed Formatting Test"
    ws1["A1"].font = Font(bold=True, size=16)

    # Merged header
    ws1.merge_cells("A1:D1")
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Various formats
    ws1["A3"] = "Bold"
    ws1["A3"].font = Font(bold=True)
    ws1["B3"] = "Italic"
    ws1["B3"].font = Font(italic=True)
    ws1["C3"] = "Red"
    ws1["C3"].font = Font(color="FF0000")
    ws1["D3"] = "Blue Fill"
    ws1["D3"].fill = PatternFill(start_color="0000FF", fill_type="solid")
    ws1["D3"].font = Font(color="FFFFFF")

    # Borders
    side = Side(style="medium", color="000000")
    for r in range(5, 8):
        for c in range(1, 5):
            cell = ws1.cell(row=r, column=c, value=f"R{r}C{c}")
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    # Merged block
    ws1.merge_cells("A9:B10")
    ws1["A9"] = "Merged Block"
    ws1["A9"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A9"].fill = PatternFill(start_color="FFFF00", fill_type="solid")

    # ── Sheet 2: Charts ──
    ws2 = wb.create_sheet("Charts")
    data = [
        ["Quarter", "Revenue", "Costs"],
        ["Q1", 100, 80],
        ["Q2", 150, 90],
        ["Q3", 130, 85],
        ["Q4", 200, 110],
    ]
    for row in data:
        ws2.append(row)

    bar = BarChart()
    bar.title = "Revenue vs Costs"
    cats = Reference(ws2, min_col=1, min_row=2, max_row=5)
    vals = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=5)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    ws2.add_chart(bar, "E2")

    line = LineChart()
    line.title = "Trend"
    line.add_data(vals, titles_from_data=True)
    line.set_categories(cats)
    ws2.add_chart(line, "E18")

    # ── Sheet 3: Drawings ──
    ws3 = wb.create_sheet("Drawings")
    ws3["A1"] = "Image and shape test sheet"

    # Add a small image
    img = PILImage.new("RGB", (60, 60), (0, 200, 0))
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    xl_img = XlImage(buf)
    xl_img.anchor = "A3"
    ws3.add_image(xl_img)

    ws3["D3"] = "[Shapes require hand-crafted XML]"

    # ── Sheet 4: Conditional ──
    ws4 = wb.create_sheet("Conditional")
    ws4["A1"] = "Color Scale"
    ws4["A1"].font = Font(bold=True)
    for i in range(2, 12):
        ws4.cell(row=i, column=1, value=(i - 1) * 10)

    ws4.conditional_formatting.add(
        "A2:A11",
        ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00FF00",
        ),
    )

    ws4["C1"] = "Data Bars"
    ws4["C1"].font = Font(bold=True)
    for i in range(2, 12):
        ws4.cell(row=i, column=3, value=(i - 1) * 10)

    ws4.conditional_formatting.add(
        "C2:C11",
        DataBarRule(start_type="min", end_type="max", color="4472C4"),
    )

    # ── Sheet 5: Interaction Grid ──
    ws5 = wb.create_sheet("Interaction")
    ws5["A1"] = "Navigation & Editing Test Grid"
    ws5["A1"].font = Font(bold=True)
    ws5.freeze_panes = "B2"

    for r in range(2, 22):
        ws5.cell(row=r, column=1, value=f"Row {r-1}")
        for c in range(2, 8):
            ws5.cell(row=r, column=c, value=(r - 1) * (c - 1))

    # Table
    ws6 = wb.create_sheet("Table")
    table_data = [
        ["Product", "Units", "Price"],
        ["Widget", 100, 9.99],
        ["Gadget", 200, 19.99],
        ["Doohickey", 50, 29.99],
    ]
    for row in table_data:
        ws6.append(row)

    tab = Table(displayName="ProductTable", ref="A1:C4")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws6.add_table(tab)

    path = out_dir / "kitchen-sink.xlsx"
    wb.save(path)
    return path
