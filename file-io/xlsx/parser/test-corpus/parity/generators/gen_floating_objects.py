"""Floating object fixtures: images, shapes, textboxes, connectors, groups."""

from pathlib import Path
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_image_png(out_dir))
    files.append(_image_jpg(out_dir))
    files.append(_shape_rectangle(out_dir))
    files.append(_shape_oval(out_dir))
    files.append(_textbox(out_dir))
    files.append(_shape_with_text(out_dir))
    return files


def _make_solid_image(color: tuple[int, int, int], size: tuple[int, int] = (80, 80), fmt: str = "PNG") -> BytesIO:
    """Create a small solid-color image in memory."""
    img = PILImage.new("RGB", size, color)
    buf = BytesIO()
    img.save(buf, format=fmt)
    buf.seek(0)
    return buf


def _image_png(out_dir: Path) -> Path:
    """Embedded PNG image anchored to cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PNG Image"
    ws["A1"] = "Image below (PNG 80x80 red)"

    img_buf = _make_solid_image((255, 0, 0), (80, 80), "PNG")
    img = XlImage(img_buf)
    img.anchor = "A3"
    ws.add_image(img)

    path = out_dir / "image-png.xlsx"
    wb.save(path)
    return path


def _image_jpg(out_dir: Path) -> Path:
    """Embedded JPEG image."""
    wb = Workbook()
    ws = wb.active
    ws.title = "JPG Image"
    ws["A1"] = "Image below (JPG 80x60 blue)"

    img_buf = _make_solid_image((0, 0, 255), (80, 60), "JPEG")
    img = XlImage(img_buf)
    img.anchor = "A3"
    ws.add_image(img)

    path = out_dir / "image-jpg.xlsx"
    wb.save(path)
    return path


def _shape_rectangle(out_dir: Path) -> Path:
    """Basic rectangle shape - placeholder (requires hand-crafted XML for true shapes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rectangle"
    ws["A1"] = "Rectangle shape below"
    ws["A3"] = "[Shape fixtures require hand-crafted XML - see README]"

    path = out_dir / "shape-rectangle.xlsx"
    wb.save(path)
    return path


def _shape_oval(out_dir: Path) -> Path:
    """Oval shape - placeholder (requires hand-crafted XML)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Oval"
    ws["A1"] = "Oval shape"
    ws["A3"] = "[Shape fixtures require hand-crafted XML - see README]"

    path = out_dir / "shape-oval.xlsx"
    wb.save(path)
    return path


def _textbox(out_dir: Path) -> Path:
    """Standalone textbox - placeholder (requires hand-crafted XML)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Textbox"
    ws["A1"] = "Textbox"
    ws["A3"] = "[Textbox fixtures require hand-crafted XML - see README]"

    path = out_dir / "textbox.xlsx"
    wb.save(path)
    return path


def _shape_with_text(out_dir: Path) -> Path:
    """Shape containing text - placeholder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shape With Text"
    ws["A1"] = "Shape with text"
    ws["A3"] = "[Shape+text fixtures require hand-crafted XML - see README]"

    path = out_dir / "shape-with-text.xlsx"
    wb.save(path)
    return path
