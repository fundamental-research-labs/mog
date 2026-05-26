"""Overlay fixtures: comments, notes."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_comments(out_dir))
    return files


def _comments(out_dir: Path) -> Path:
    """Cell comments/notes (classic yellow sticky notes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    # Cell with a short comment
    ws["A1"] = "Cell with comment"
    ws["A1"].comment = Comment("This is a simple comment.", "Test Author")

    # Cell with a longer multi-line comment
    ws["A2"] = "Multi-line comment"
    ws["A2"].comment = Comment(
        "Line 1: This comment has multiple lines.\n"
        "Line 2: It should preserve newlines.\n"
        "Line 3: And display them correctly.",
        "Test Author",
    )

    # Cell with comment but no value
    ws["A3"].comment = Comment("Comment on empty cell", "Another Author")

    # Cell with formatted value and comment
    ws["B1"] = 42
    ws["B1"].comment = Comment("The answer to everything.", "Deep Thought")

    # Cell with a very long comment
    ws["B2"] = "Long comment"
    ws["B2"].comment = Comment(
        "This is a very long comment that tests how the renderer handles "
        "overflow text in the comment popup. It should wrap properly and "
        "not clip or truncate the content. The comment box should resize "
        "to accommodate all this text or provide scrolling.",
        "Verbose Author",
    )

    # Multiple comments in a column
    for i in range(1, 6):
        ws.cell(row=i, column=3, value=f"Item {i}")
        ws.cell(row=i, column=3).comment = Comment(
            f"Note for item {i}", "Batch Author"
        )

    path = out_dir / "comments.xlsx"
    wb.save(str(path))
    return path
