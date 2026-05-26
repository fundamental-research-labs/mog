"""Pivot table fixture: source data + pivot table XML injected into XLSX zip.

openpyxl cannot create pivot tables from scratch, so we:
1. Create the workbook with source data and pre-computed pivot output using openpyxl
2. Save the XLSX
3. Re-open the zip and inject the pivot table XML parts (cache def, cache records, pivot table)
4. Update Content_Types, workbook.xml, and relationship files
"""

import shutil
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font


def generate(out_dir: Path) -> list[Path]:
    files = []
    files.append(_basic_pivot(out_dir))
    return files


# ---------------------------------------------------------------------------
# XML namespace map for OOXML
# ---------------------------------------------------------------------------
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

REL_PIVOT_CACHE_DEF = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"
)
REL_PIVOT_TABLE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"
)
CT_PIVOT_TABLE = "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"
CT_PIVOT_CACHE_DEF = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"
)
CT_PIVOT_CACHE_RECORDS = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"
)


def _basic_pivot(out_dir: Path) -> Path:
    """Create a minimal pivot table: Category x Region -> Sum(Amount).

    Source data (Sheet1 "Data"):
        Category | Region | Amount
        Fruit    | East   | 100
        Fruit    | West   | 150
        Veggie   | East   | 200
        Veggie   | West   | 250

    Pivot output (Sheet2 "Pivot"):
        Row Labels | East | West | Grand Total
        Fruit      | 100  | 150  | 250
        Veggie     | 200  | 250  | 450
        Grand Total| 300  | 400  | 700
    """
    wb = Workbook()

    # --- Sheet1: Source data ---
    ws_data = wb.active
    ws_data.title = "Data"
    headers = ["Category", "Region", "Amount"]
    rows = [
        ("Fruit", "East", 100),
        ("Fruit", "West", 150),
        ("Veggie", "East", 200),
        ("Veggie", "West", 250),
    ]
    ws_data.append(headers)
    for row in rows:
        ws_data.append(row)

    # --- Sheet2: Pre-computed pivot output (matches what Excel would cache) ---
    ws_pivot = wb.create_sheet("Pivot")

    # Header row
    pivot_headers = ["", "East", "West", "Grand Total"]
    for col_idx, val in enumerate(pivot_headers, 1):
        cell = ws_pivot.cell(row=1, column=col_idx, value=val)
        cell.font = Font(bold=True)

    # Data rows
    pivot_rows = [
        ("Fruit", 100, 150, 250),
        ("Veggie", 200, 250, 450),
        ("Grand Total", 300, 400, 700),
    ]
    for row_idx, row in enumerate(pivot_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws_pivot.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 4 or col_idx == 1:
                cell.font = Font(bold=True)

    # Save base workbook
    path = out_dir / "pivot-basic.xlsx"
    wb.save(str(path))

    # --- Inject pivot table XML into the XLSX zip ---
    _inject_pivot_xml(path)

    return path


def _inject_pivot_xml(xlsx_path: Path) -> None:
    """Open the XLSX zip and inject pivot table/cache XML parts."""
    tmp = xlsx_path.with_suffix(".tmp.xlsx")

    with ZipFile(str(xlsx_path), "r") as zin, ZipFile(str(tmp), "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == "[Content_Types].xml":
                data = _patch_content_types(data)
            elif item.filename == "xl/_rels/workbook.xml.rels":
                data = _patch_workbook_rels(data)
            elif item.filename == "xl/workbook.xml":
                data = _patch_workbook_xml(data)

            zout.writestr(item, data)

        # Add pivot cache parts
        zout.writestr("xl/pivotCache/pivotCacheDefinition1.xml", _pivot_cache_definition_xml())
        zout.writestr("xl/pivotCache/pivotCacheRecords1.xml", _pivot_cache_records_xml())
        zout.writestr(
            "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
            _pivot_cache_rels_xml(),
        )
        # Add pivot table
        zout.writestr("xl/pivotTables/pivotTable1.xml", _pivot_table_xml())

        # Add sheet2 rels (sheet -> pivot table)
        sheet2_rels_path = "xl/worksheets/_rels/sheet2.xml.rels"
        existing_rels = None
        try:
            existing_rels = zin.read(sheet2_rels_path)
        except KeyError:
            pass
        zout.writestr(sheet2_rels_path, _sheet2_rels_xml(existing_rels))

    shutil.move(str(tmp), str(xlsx_path))


def _patch_content_types(data: bytes) -> bytes:
    """Add pivot table content type overrides."""
    ET.register_namespace("", NS_CT)
    root = ET.fromstring(data)

    overrides = [
        ("/xl/pivotTables/pivotTable1.xml", CT_PIVOT_TABLE),
        ("/xl/pivotCache/pivotCacheDefinition1.xml", CT_PIVOT_CACHE_DEF),
        ("/xl/pivotCache/pivotCacheRecords1.xml", CT_PIVOT_CACHE_RECORDS),
    ]
    for part_name, content_type in overrides:
        elem = ET.SubElement(root, f"{{{NS_CT}}}Override")
        elem.set("PartName", part_name)
        elem.set("ContentType", content_type)

    return ET.tostring(root, xml_declaration=True, encoding="UTF-8")


def _patch_workbook_rels(data: bytes) -> bytes:
    """Add pivot cache relationship to workbook rels."""
    ET.register_namespace("", NS_REL)
    root = ET.fromstring(data)

    rel = ET.SubElement(root, f"{{{NS_REL}}}Relationship")
    rel.set("Id", "rIdPivotCache1")
    rel.set("Type", REL_PIVOT_CACHE_DEF)
    rel.set("Target", "pivotCache/pivotCacheDefinition1.xml")

    return ET.tostring(root, xml_declaration=True, encoding="UTF-8")


def _patch_workbook_xml(data: bytes) -> bytes:
    """Add pivotCaches element to workbook.xml."""
    ET.register_namespace("", NS)
    ET.register_namespace("r", NS_R)

    root = ET.fromstring(data)

    pivot_caches = ET.SubElement(root, f"{{{NS}}}pivotCaches")
    pivot_cache = ET.SubElement(pivot_caches, f"{{{NS}}}pivotCache")
    pivot_cache.set("cacheId", "1")
    pivot_cache.set(f"{{{NS_R}}}id", "rIdPivotCache1")

    return ET.tostring(root, xml_declaration=True, encoding="UTF-8")


def _pivot_cache_definition_xml() -> bytes:
    """Minimal pivot cache definition XML."""
    xml = f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="{NS}"
    xmlns:r="{NS_R}"
    r:id="rId1"
    refreshOnLoad="0"
    recordCount="4">
  <cacheSource type="worksheet">
    <worksheetSource ref="A1:C5" sheet="Data"/>
  </cacheSource>
  <cacheFields count="3">
    <cacheField name="Category" numFmtId="0">
      <sharedItems count="2">
        <s v="Fruit"/>
        <s v="Veggie"/>
      </sharedItems>
    </cacheField>
    <cacheField name="Region" numFmtId="0">
      <sharedItems count="2">
        <s v="East"/>
        <s v="West"/>
      </sharedItems>
    </cacheField>
    <cacheField name="Amount" numFmtId="0">
      <sharedItems containsSemiMixedTypes="0" containsString="0"
                   containsNumber="1" minValue="100" maxValue="250"/>
    </cacheField>
  </cacheFields>
</pivotCacheDefinition>"""
    return xml.encode("utf-8")


def _pivot_cache_records_xml() -> bytes:
    """Pivot cache records — the raw source data rows."""
    xml = f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheRecords xmlns="{NS}"
    xmlns:r="{NS_R}"
    count="4">
  <r><x v="0"/><x v="0"/><n v="100"/></r>
  <r><x v="0"/><x v="1"/><n v="150"/></r>
  <r><x v="1"/><x v="0"/><n v="200"/></r>
  <r><x v="1"/><x v="1"/><n v="250"/></r>
</pivotCacheRecords>"""
    return xml.encode("utf-8")


def _pivot_cache_rels_xml() -> bytes:
    """Rels for pivot cache definition -> cache records."""
    xml = f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords"
    Target="pivotCacheRecords1.xml"/>
</Relationships>"""
    return xml.encode("utf-8")


def _pivot_table_xml() -> bytes:
    """Minimal pivot table definition XML.

    Layout: Category on rows, Region on columns, Sum(Amount) as data.
    Output location: A1 on the Pivot sheet.
    """
    xml = f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="{NS}"
    name="PivotTable1"
    cacheId="1"
    dataOnRows="0"
    applyNumberFormats="0"
    applyBorderFormats="0"
    applyFontFormats="0"
    applyPatternFormats="0"
    applyAlignmentFormats="0"
    applyWidthHeightFormats="1"
    dataCaption="Values"
    grandTotalCaption="Grand Total"
    showDrill="1"
    showDataTips="1"
    useAutoFormatting="1"
    rowGrandTotals="1"
    colGrandTotals="1"
    itemPrintTitles="1"
    createdVersion="6"
    indent="0"
    outline="1"
    outlineData="1"
    multipleFieldFilters="0">
  <location ref="A1:D4" firstHeaderRow="1" firstDataRow="1" firstDataCol="1"
            rowPageCount="0" colPageCount="0"/>
  <pivotFields count="3">
    <pivotField axis="axisRow" showAll="0">
      <items count="3">
        <item x="0"/>
        <item x="1"/>
        <item t="default"/>
      </items>
    </pivotField>
    <pivotField axis="axisCol" showAll="0">
      <items count="3">
        <item x="0"/>
        <item x="1"/>
        <item t="default"/>
      </items>
    </pivotField>
    <pivotField dataField="1" showAll="0"/>
  </pivotFields>
  <rowFields count="1">
    <field x="0"/>
  </rowFields>
  <rowItems count="3">
    <i><x/></i>
    <i><x v="1"/></i>
    <i t="grand"><x/></i>
  </rowItems>
  <colFields count="1">
    <field x="1"/>
  </colFields>
  <colItems count="3">
    <i><x/></i>
    <i><x v="1"/></i>
    <i t="grand"><x/></i>
  </colItems>
  <dataFields count="1">
    <dataField name="Sum of Amount" fld="2" subtotal="sum"
               baseField="0" baseItem="0"/>
  </dataFields>
</pivotTableDefinition>"""
    return xml.encode("utf-8")


def _sheet2_rels_xml(existing: bytes | None) -> bytes:
    """Rels for sheet2 -> pivot table. Merges with existing rels if present."""
    if existing:
        ET.register_namespace("", NS_REL)
        root = ET.fromstring(existing)
    else:
        root = ET.Element(f"{{{NS_REL}}}Relationships")

    rel = ET.SubElement(root, f"{{{NS_REL}}}Relationship")
    rel.set("Id", "rIdPT1")
    rel.set("Type", REL_PIVOT_TABLE)
    rel.set("Target", "../pivotTables/pivotTable1.xml")

    return ET.tostring(root, xml_declaration=True, encoding="UTF-8")
